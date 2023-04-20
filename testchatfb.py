
import unittest
import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from unidecode import unidecode
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


class TestFacebook(unittest.TestCase):
    def setUp(self):
        self.browser = webdriver.Edge()

    def tearDown(self):
        self.browser.quit()

    def read_excel(self, case):
        if(case == "taikhoan"):
            taikhoan =""
            matkhau =""
            dataframe = openpyxl.load_workbook("excel/Book2.xlsx")
            dataframe1 = dataframe['Chat']
            
            taikhoan = dataframe1[f'B3'].value
            matkhau = dataframe1[f'C3'].value
            return taikhoan, matkhau
        if(case == "search"):
            ten=""
            dataframe = openpyxl.load_workbook("Book2.xlsx")
            dataframe1 = dataframe['Chat']
            ten = dataframe1[f'B5'].value
            return ten
        if(case == "text"):
            id=""
            ten=""
            noidung=""
            dataframe = openpyxl.load_workbook("Book2.xlsx")
            dataframe1 = dataframe['Chat']
            id = dataframe1[f'B7'].value
            ten = dataframe1[f'C7'].value
            noidung = dataframe1[f'D7'].value
            return id,ten,noidung
        if(case == "image"):
            id=""
            ten=""
            noidung=""
            dataframe = openpyxl.load_workbook("Book2.xlsx")
            dataframe1 = dataframe['Chat']
            id = dataframe1[f'B9'].value
            ten = dataframe1[f'C9'].value
            noidung = dataframe1[f'D9'].value
            return id,ten,noidung
        if(case == "like"):
            id=""
            ten=""
            dataframe = openpyxl.load_workbook("Book2.xlsx")
            dataframe1 = dataframe['Chat']
            id = dataframe1[f'B11'].value
            ten = dataframe1[f'C11'].value
            return id,ten
        if(case == "emoji"):
            id=""
            ten=""
            icon=""
            dataframe = openpyxl.load_workbook("Book2.xlsx")
            dataframe1 = dataframe['Chat']
            id = dataframe1[f'B13'].value
            ten = dataframe1[f'C13'].value
            icon = dataframe1[f'D13'].value
            return id,ten,icon
        if(case == "theme"):
            id=""
            ten=""
            idTheme=""
            dataframe = openpyxl.load_workbook("Book2.xlsx")
            dataframe1 = dataframe['Chat']
            id = dataframe1[f'B15'].value
            ten = dataframe1[f'C15'].value
            idTheme = dataframe1[f'D15'].value
            return id,ten,idTheme
        if(case == "iconMessage"):
            id=""
            ten=""
            icon=""
            noidung=""
            dataframe = openpyxl.load_workbook("Book2.xlsx")
            dataframe1 = dataframe['Chat']
            id = dataframe1[f'B17'].value
            ten = dataframe1[f'C17'].value
            icon = dataframe1[f'D17'].value
            noidung = dataframe1[f'E17'].value
            return id,ten,icon,noidung
        if(case == "sticker"):
            id=""
            ten=""
            type=""
            sticker=""
            box_check=""
            dataframe = openpyxl.load_workbook("Book2.xlsx")
            dataframe1 = dataframe['Chat']
            id = dataframe1[f'B19'].value
            ten = dataframe1[f'C19'].value
            type = dataframe1[f'D19'].value
            sticker = dataframe1[f'E19'].value
            box_check = dataframe1[f'F19'].value
            return id,ten,type,sticker,box_check

    def login_open_chat(self):
        excel = self.read_excel("taikhoan")
        self.browser.get("https://www.facebook.com")
        self.browser.maximize_window()

        # login
        user_name = self.browser.find_element(By.ID, "email").send_keys(excel[0])
        pass_word = self.browser.find_element(By.ID, "pass").send_keys(excel[1])
        btn_login = self.browser.find_element(By.NAME, "login").click()

        time.sleep(5)
        self.browser.find_element(By.XPATH, '//*[@id="facebook"]/body').click()

        # navigate to chat and search friend
        self.browser.get("https://www.facebook.com/messages/t/100010705752704/")
        time.sleep(5)
        self.browser.find_element(By.XPATH, '//*[@id="facebook"]/body').click()
    
    def search_friend(self, name):
        search_chat = self.browser.find_element(By.XPATH, "//input[@aria-label='Tìm kiếm trên Messenger']")
        time.sleep(2)
        search_chat.click()
        search_chat.send_keys(name)

    def test_search(self):
        # login and open chat
        self.login_open_chat()
        time.sleep(5)
        excel = self.read_excel("search")
        name = excel[0]
        self.search_friend(name)
        
        #check friend is display
        time.sleep(2)
        list_friend = self.browser.find_element(By.XPATH, "//li[@class='xexx8yu xsyo7zv x18d9i69 x16hj40l']")
        friends = list_friend.find_elements(By.XPATH, ".//span[@class='x1lliihq x6ikm8r x10wlt62 x1n2onr6 xlyipyv xuxw1ft']")
        for friend in friends:

            self.assertIn(name, unidecode(friend.text.lower()))

        time.sleep(5)

    def test_login_send_text_message(self):
        excel = self.read_excel("text")
        id=excel[0]
        ten=excel[1]
        noidung=excel[2]

        # login and open chat
        self.login_open_chat()
        time.sleep(5)
        self.search_friend(ten)

        # select friend
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//li[@id='{id}']").click()

        # find all message = a 
        time.sleep(2)
        box_message = self.browser.find_elements(By.XPATH, f"//*[contains(text(), '{noidung}')]")
        time.sleep(2)
        len_message_before = len(box_message)

        # send text message
        time.sleep(2)
        input_message = self.browser.find_element(By.XPATH, "//div[@aria-label='Tin nhắn']")      
        time.sleep(2)
        input_message.click()
        time.sleep(2)
        input_message.send_keys(excel[2])
        time.sleep(2)
        input_message.send_keys(Keys.ENTER)

        # check message is display
        time.sleep(2)
        box_message_2 = self.browser.find_elements(By.XPATH, f"//*[contains(text(), '{noidung}')]")
        print(box_message_2)
        time.sleep(2)
        len_message_after = len(box_message_2)
        print(f'after :{len_message_after} ')

        self.assertLess(len_message_before, len_message_after, "text upload failed")
        
        time.sleep(5)
        
    def test_login_send_image_message(self):
        excel = self.read_excel("image")
        id=excel[0]
        ten=excel[1]
        noidung=excel[2]

        # login and open chat
        self.login_open_chat()
        time.sleep(5)
        self.search_friend(ten)

        # select friend
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//li[@id='{id}']").click()

        # find all img
        time.sleep(2)
        box_img = self.browser.find_elements(By.XPATH, "//img[@alt='Mở ảnh']")
        time.sleep(2)
        len_img_before = len(box_img)

        #send image
        time.sleep(2)
        input_image = self.browser.find_element(By.XPATH, "//input[@type='file']")      
        time.sleep(2)
        input_image.send_keys(noidung)
        time.sleep(2)
        input_message = self.browser.find_element(By.XPATH, "//div[@aria-label='Tin nhắn']")
        time.sleep(2)
        input_message.send_keys(Keys.ENTER)

        # check img is display
        time.sleep(2)
        box_img_2 = self.browser.find_elements(By.XPATH, "//img[@alt='Mở ảnh']")
        time.sleep(2)
        len_img_after = len(box_img_2)
        self.assertLess(len_img_before, len_img_after, "Image upload failed")
        time.sleep(5)
    
    def test_login_send_like_icon_message(self):
        excel = self.read_excel("like")
        id=excel[0]
        ten=excel[1]
        # login and open chat
        self.login_open_chat()
        time.sleep(5)
        self.search_friend(ten)

        # select friend
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//li[@id='{id}']").click()

        # find all icon
        time.sleep(2)
        box_icon = self.browser.find_elements(By.XPATH, "//span[@class='x3nfvp2 x1j61x8r x1fcty0u xdj266r xhhsvwb xat24cr xgzva0m xxymvpz x10w6t97 x1td3qas']")
        time.sleep(2)
        len_icon_before = len(box_icon)

        # send like
        button_like = self.browser.find_element(By.XPATH, "//div[@class='x1i10hfl x1qjc9v5 xjbqb8w xjqpnuy xa49m3k xqeqjp1 x2hbi6w x13fuv20 xu3j5b3 x1q0q8m5 x26u7qi x972fbf xcfux6l x1qhh985 xm0m39n x9f619 x1ypdohk xdl72j9 x2lah0s xe8uvvx xdj266r xat24cr x2lwn1j xeuugli x1n2onr6 x16tdsg8 x1hl2dhg xggy1nq x1ja2u2z x1t137rt x1o1ewxj x3x9cwd x1e5q0jg x13rtm0m x3nfvp2 x1q0g3np x87ps6o x1lku1pv x1a2a7pz xsgj6o6 xw3qccf xcu9agk x2qib4z x1y1aw1k xwib8y2']")
        time.sleep(2)
        button_like.click()

        # check like is display
        time.sleep(2)
        box_icon_2 = self.browser.find_elements(By.XPATH, "//span[@class='x3nfvp2 x1j61x8r x1fcty0u xdj266r xhhsvwb xat24cr xgzva0m xxymvpz x10w6t97 x1td3qas']")
        time.sleep(2)
        len_icon_after = len(box_icon_2)
        self.assertLess(len_icon_before, len_icon_after, "Like icon upload failed")
        time.sleep(5)

    def test_login_send_emoji_message(self):
        excel = self.read_excel("emoji")
        id=excel[0]
        ten=excel[1]
        icon=excel[2]
        # login and open chat
        self.login_open_chat()
        time.sleep(5)
        self.search_friend(ten)
        # select friend
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//li[@id='{id}']").click()
        # find all icon
        time.sleep(2)
        box_icon = self.browser.find_elements(By.XPATH, "//span[@class='x3nfvp2 x1j61x8r x1fcty0u xdj266r xhhsvwb xat24cr xgzva0m xxymvpz x10w6t97 x1td3qas']")
        time.sleep(2)
        len_icon_before = len(box_icon)

        # send emoji
        button_icon = self.browser.find_element(By.XPATH, "//div[@aria-label='Chọn biểu tượng cảm xúc']")
        time.sleep(2)
        button_icon.click()
        time.sleep(5)
        icon = self.browser.find_element(By.XPATH, f"//img[@src='{icon}']")
        time.sleep(2)
        icon.click()
        time.sleep(5)
        input_message = self.browser.find_element(By.XPATH, "//div[@aria-label='Tin nhắn']")
        time.sleep(2)
        input_message.send_keys(Keys.ENTER)

        # check emoji is display
        time.sleep(2)
        box_icon_2 = self.browser.find_elements(By.XPATH, "//span[@class='x3nfvp2 x1j61x8r x1fcty0u xdj266r xhhsvwb xat24cr xgzva0m xxymvpz x10w6t97 x1td3qas']")
        time.sleep(2)
        len_icon_after = len(box_icon_2)
        self.assertLess(len_icon_before, len_icon_after, "Emoji upload failed")
        time.sleep(5)
    
    def test_login_change_theme_chat(self):
        excel = self.read_excel("change_theme")
        id = excel[0]
        ten = excel[1]
        idTheme = excel[2]
        list_theme = ['737761000603635', '262191918210707', '788274591712841', '275041734441112']
        # login and open chat
        self.login_open_chat()
        time.sleep(5)
        self.search_friend(ten)

        # select friend
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//li[@id='{id}']").click()

        # find notify change theme
        time.sleep(2)
        notify_change = self.browser.find_elements(By.XPATH, "//div[contains(text(), 'Bạn đã đổi chủ đề')]")
        time.sleep(2)
        box_notify_change_before = len(notify_change)

        # change theme
        self.browser.find_element(By.XPATH, "//*[contains(text(), 'Tùy chỉnh đoạn chat')]").click()
        time.sleep(2)
        self.browser.find_element(By.XPATH, "//*[contains(text(), 'Đổi chủ đề')]").click()
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//div[@id='{idTheme}']").click()
        time.sleep(2)
        element = WebDriverWait(self.browser, 10).until(
        EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Lưu'and @tabindex='0']"))
        )
        element.click()

        # check change theme
        time.sleep(2)
        notify_change_2 = self.browser.find_elements(By.XPATH, "//div[contains(text(), 'Bạn đã đổi chủ đề')]")
        time.sleep(2)
        box_notify_change_after = len(notify_change_2)
        self.assertLess(box_notify_change_before, box_notify_change_after, "Change theme failed")
        time.sleep(5)

    def test_emoji_message(self):
        excel = self.read_excel("iconMessage")
        id = excel[0]
        ten = excel[1]
        icon = excel[2]
        noidung = excel[3]
        # login and open chat
        self.login_open_chat()
        time.sleep(5)
        self.search_friend(ten)
        # select friend
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//li[@id='{id}']").click()
        # find all icon
        time.sleep(2)
        box_icon = self.browser.find_elements(By.XPATH, "//span[@class='x3nfvp2 x1j61x8r x1fcty0u xdj266r xhhsvwb xat24cr xgzva0m xxymvpz x10w6t97 x1td3qas']")
        time.sleep(2)
        len_icon_before = len(box_icon)

        # send emoji
        uk_message = self.browser.find_element(By.XPATH, f"//*[contains(text(), '{noidung}')]")
        actions = ActionChains(self.browser)
        actions.move_to_element(uk_message).perform()
        time.sleep(2)
        element = WebDriverWait(self.browser, 10).until(
        EC.presence_of_element_located((By.XPATH, '//div[@aria-label="Bày tỏ cảm xúc"]'))
        )
        element.click()
        time.sleep(2)
        aggry_icon = self.browser.find_element(By.XPATH, f"//img[@src='{icon}']")
        time.sleep(2)
        aggry_icon.click()

        # check emoji is display
        time.sleep(2)
        box_icon_2 = self.browser.find_elements(By.XPATH, "//span[@class='x3nfvp2 x1j61x8r x1fcty0u xdj266r xhhsvwb xat24cr xgzva0m xxymvpz x10w6t97 x1td3qas']")
        time.sleep(2)
        len_icon_after = len(box_icon_2)
        self.assertLess(len_icon_before, len_icon_after, "Emoji on message failed")
        time.sleep(5)

    def test_send_sticker(self):
        excel = self.read_excel("sticker")
        id = excel[0]
        ten = excel[1]
        type_sticker = excel[2]
        sticker = excel[3]
        box_check = excel[4]
        # login and open chat
        self.login_open_chat()
        time.sleep(5)
        self.search_friend(ten)

        # select friend
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//li[@id='{id}']").click()

        # find all icon
        time.sleep(2)
        box_sticker = self.browser.find_elements(By.XPATH, f"//div[@aria-label='{box_check}']")
        time.sleep(2)
        len_sticker_before = len(box_sticker)

        # send sticker
        self.browser.find_element(By.XPATH, "//div[@aria-label='Chọn nhãn dán']").click()
        time.sleep(2)
        input_sticker = self.browser.find_element(By.XPATH, "//input[@aria-label='Tìm kiếm nhãn dán']")
        input_sticker.click()
        input_sticker.send_keys(type_sticker)
        time.sleep(2)
        self.browser.find_element(By.XPATH, f"//div[@aria-label='{sticker}']").click()
        
        # check sticker is display
        time.sleep(5)
        box_sticker_2 = self.browser.find_elements(By.XPATH, f"//div[@aria-label='{box_check}']")
        time.sleep(2)
        len_sticker_after = len(box_sticker_2)
        self.assertLess(len_sticker_before, len_sticker_after, "Send sticker failed")
        time.sleep(5)


if __name__ == '__main__':
    unittest.main()